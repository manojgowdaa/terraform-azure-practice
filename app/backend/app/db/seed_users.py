"""
Seed users from users.xlsx into the database.

Usage:
    cd app/backend
    python -m app.db.seed_users

- Reads users.xlsx from the backend folder
- Hashes passwords with bcrypt
- Inserts new users, updates existing ones (by email)
- Only users in this Excel file will have login access
"""
import os
import sys
from pathlib import Path

import openpyxl
import bcrypt
from sqlalchemy.orm import Session

from app.db.database import engine, SessionLocal, Base
from app.db.models import User


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

# Path to the Excel file
EXCEL_PATH = Path(__file__).resolve().parent.parent.parent / "users.xlsx"


def seed_users():
    # Create tables if they don't exist
    Base.metadata.create_all(bind=engine)

    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH} not found!")
        print("Create users.xlsx with columns: name, email, password, role")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Read header row
    headers = [cell.value.lower().strip() for cell in ws[1]]
    print(f"Columns found: {headers}")

    db: Session = SessionLocal()
    added = 0
    updated = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        if not data.get("email") or not data.get("password"):
            continue  # skip empty rows

        name = str(data["name"]).strip()
        email = str(data["email"]).strip().lower()
        password = str(data["password"]).strip()
        role = str(data.get("role", "user")).strip().lower()

        # Check if user already exists
        existing = db.query(User).filter(User.email == email).first()

        if existing:
            # Update name, password, role
            existing.name = name
            existing.password_hash = hash_password(password)
            existing.role = role
            updated += 1
            print(f"  Updated: {email} (role={role})")
        else:
            # Create new user
            user = User(
                name=name,
                email=email,
                password_hash=hash_password(password),
                role=role,
            )
            db.add(user)
            added += 1
            print(f"  Added: {email} (role={role})")

    db.commit()
    db.close()

    print(f"\nDone! Added: {added}, Updated: {updated}")
    print("Only these users can login to the app.")


if __name__ == "__main__":
    seed_users()
